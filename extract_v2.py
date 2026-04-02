# -*- coding: utf-8 -*-
import openpyxl
import json
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def safe_str(v):
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        return str(v)
    s = str(v)
    if len(s) > 100:
        return s[:100] + '...'
    return s

def get_files(dir_path, extensions=None):
    """Get files from directory with their actual filenames"""
    if not os.path.exists(dir_path):
        return []
    result = []
    for f in os.listdir(dir_path):
        full = os.path.join(dir_path, f)
        if os.path.isfile(full):
            if extensions is None or any(f.lower().endswith(ext) for ext in extensions):
                result.append(full)
    return result

def extract_xlsx(filepath, max_rows=60, max_sheets=3):
    """Extract data from xlsx file"""
    result = {'file': os.path.basename(filepath), 'path': filepath, 'sheets': []}
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        result['sheet_names'] = wb.sheetnames
        
        for i, sheet_name in enumerate(wb.sheetnames[:max_sheets]):
            ws = wb[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'dimensions': ws.dimensions,
                'max_row': min(ws.max_row, max_rows * 2),
                'max_col': ws.max_column,
                'rows': []
            }
            
            for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), 1):
                non_none = [v for v in row if v is not None]
                if non_none:
                    sheet_data['rows'].append({
                        'row': row_idx,
                        'cells': [safe_str(v) for v in row[:25]]
                    })
            
            result['sheets'].append(sheet_data)
            
    except Exception as e:
        result['error'] = str(e)
    
    return result

# Get actual filenames from directories
budget2024_files = get_files(r'E:\2024预算', ['.xlsx', '.xls'])
budget2023_files = get_files(r'E:\2023年预算工作', ['.xlsx', '.xls', '.xlsm'])
desk_2023budget_files = [f for f in get_files(r'E:\桌面', ['.xlsx']) if '预算' in f or '预算' in os.path.basename(f)]
op_data_files = get_files(r'E:\主要经营数据', ['.xlsx'])
pl_file = get_files(r'E:\桌面', ['.xlsx'])
kpi_files = get_files(r'E:\分管部门', ['.xlsx'])
perf_file = get_files(r'E:\述职报告', ['.xlsx'])

all_files = (budget2024_files + budget2023_files + desk_2023budget_files + 
             op_data_files + pl_file + kpi_files + perf_file)

# Deduplicate
seen = set()
unique_files = []
for f in all_files:
    if f not in seen:
        seen.add(f)
        unique_files.append(f)

print(f'Found {len(unique_files)} unique files to process')

all_results = []
for f in unique_files:
    fname = os.path.basename(f)
    if any(x in fname for x in ['预算', '经营数据', 'P&L', 'KPI', '考核', '任务', '价格']):
        print(f"Processing: {fname[:60]}", flush=True)
        r = extract_xlsx(f)
        all_results.append(r)

# Write to JSON file
output_file = r"C:\Users\ericz\.openclaw\workspace\excel_data_v2.json"
with open(output_file, 'w', encoding='utf-8') as fp:
    json.dump(all_results, fp, ensure_ascii=False, indent=2)

print(f"\nWritten to {output_file}", flush=True)
print(f"Total files processed: {len(all_results)}", flush=True)
