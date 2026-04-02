# -*- coding: utf-8 -*-
import openpyxl
import json
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def safe_str(v):
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        return str(v)
    s = str(v)
    if len(s) > 80:
        return s[:80] + '...'
    return s

def extract_xlsx(filepath, max_rows=50, max_sheets=5):
    """Extract data from xlsx file"""
    result = {'file': os.path.basename(filepath), 'sheets': []}
    
    if not os.path.exists(filepath):
        result['error'] = 'FILE NOT FOUND'
        return result
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        result['sheet_names'] = wb.sheetnames
        
        for i, sheet_name in enumerate(wb.sheetnames[:max_sheets]):
            if i >= max_sheets:
                break
            ws = wb[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'dimensions': ws.dimensions,
                'max_row': ws.max_row,
                'max_col': ws.max_column,
                'rows': []
            }
            
            for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), 1):
                non_none = [v for v in row if v is not None]
                if non_none:
                    sheet_data['rows'].append({
                        'row': row_idx,
                        'cells': [safe_str(v) for v in row[:20]]
                    })
            
            result['sheets'].append(sheet_data)
            
    except Exception as e:
        result['error'] = str(e)
    
    return result

# Files to extract
files = [
    r"E:\2024预算\附件5. 重庆丽苑酒店2024年酒店经营预算模板-汇总版V1.7-0129自用版.xlsx",
    r"E:\2023年预算工作\附件5. 重庆丽苑酒店2023年酒店经营预算模板-汇总版V3.3.4-1207（二上版基础上增加事业群分摊费用267466元+净利润增加60万元）.xlsx",
    r"E:\主要经营数据\主要经营数据1-4月.xlsx",
    r"E:\桌面\酒店经营月报P&L 202212 重庆丽苑.xlsx",
    r"E:\述职报告\2023年度经营业绩考核完成情况（干部） - 张实(2).xlsx",
    r"E:\分管部门\2024部门KPI（餐饮，客房，前厅，营销）(5月).xlsx",
    r"E:\桌面\2023年月度任务划分汇总（2643万）.xlsx",
    r"E:\桌面\附件5. 重庆丽苑酒店2023年酒店经营预算模板-汇总版V3.0-1008二稿调整板.xlsx",
]

all_results = []
for f in files:
    print(f"Processing: {os.path.basename(f)}", flush=True)
    r = extract_xlsx(f)
    all_results.append(r)

# Write to JSON file
output_file = r"C:\Users\ericz\.openclaw\workspace\excel_data.json"
with open(output_file, 'w', encoding='utf-8') as fp:
    json.dump(all_results, fp, ensure_ascii=False, indent=2)

print(f"\nWritten to {output_file}", flush=True)
print(f"Total files processed: {len(all_results)}", flush=True)
